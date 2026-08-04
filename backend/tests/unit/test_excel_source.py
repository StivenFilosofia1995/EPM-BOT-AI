"""Tests de `XlsxProgramacionSource` contra el archivo real.

El criterio de aceptación de P2A vive aquí. No usa red ni base de datos: lee el
`.xlsx` que está en el repositorio, que es el que produjo la Fundación en julio
de 2026.

Si estos tests fallan tras un cambio, lo que se rompió es la interpretación de
la parrilla, y eso llega directo a lo que el bot le dice a la gente.
"""

from datetime import UTC, datetime
from pathlib import Path
from typing import Any, cast
from uuid import uuid4
from zoneinfo import ZoneInfo

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.application.ingestion.schemas import IngestionWarning, RowStatus
from src.infrastructure.ingestion.excel.rooms import RoomCatalog
from src.infrastructure.ingestion.excel.source import ImportContext, XlsxProgramacionSource

BOGOTA = ZoneInfo("America/Bogota")

SAMPLE = (
    Path(__file__).resolve().parents[3] / "Programacion_Formativa Biblioteca_Julio_2026.xlsx"
)

#: Las salas reales de Biblioteca EPM (contrato §5).
ROOM_NAMES = (
    "Taller Infantil",
    "Sala de formación",
    "Sala de Formación 3",
    "Sala de formación 4",
    "Sala de Investigadores",
    "Cafetería",
)

pytestmark = pytest.mark.skipif(
    not SAMPLE.exists(), reason=f"Falta el archivo de muestra: {SAMPLE.name}"
)


@pytest.fixture
def catalog() -> RoomCatalog:
    return RoomCatalog.from_pairs([(name, uuid4()) for name in ROOM_NAMES])


@pytest.fixture
def context(catalog: RoomCatalog) -> ImportContext:
    return ImportContext(
        venue_slug="biblioteca-epm",
        year=2026,
        month=7,
        rooms=catalog,
        file_name=SAMPLE.name,
    )


@pytest.fixture
def report(context: ImportContext):  # type: ignore[no-untyped-def]
    return XlsxProgramacionSource().parse(SAMPLE, context)


class TestAcceptanceCriteria:
    """El criterio de aceptación de P2A, verificado contra el archivo real."""

    def test_reads_23_data_rows(self, report) -> None:  # type: ignore[no-untyped-def]
        assert report.rows_read == 23

    def test_produces_50_activities(self, report) -> None:  # type: ignore[no-untyped-def]
        """23 filas → 50 actividades tras expandir las fechas.

        23 son FILAS, no actividades: cada fila con N fechas produce N.
        """
        assert len(report.extractions) == 50

    def test_breakdown_by_sheet(self, report) -> None:  # type: ignore[no-untyped-def]
        by_sheet: dict[str, int] = {}
        rows_by_sheet: dict[str, int] = {}
        for row in report.rows:
            by_sheet[row.sheet] = by_sheet.get(row.sheet, 0) + len(row.extractions)
            rows_by_sheet[row.sheet] = rows_by_sheet.get(row.sheet, 0) + 1

        infantil = next(s for s in by_sheet if "infantil" in s.lower())
        adultos = next(s for s in by_sheet if "adultos" in s.lower())

        assert rows_by_sheet[infantil] == 8
        assert by_sheet[infantil] == 11
        assert rows_by_sheet[adultos] == 15
        assert by_sheet[adultos] == 39

    def test_no_row_errors(self, report) -> None:  # type: ignore[no-untyped-def]
        errors = [f"{r.location}: {r.errors}" for r in report.rows if r.status is RowStatus.ERROR]
        assert not errors, errors

    def test_every_activity_has_evidence(self, report) -> None:  # type: ignore[no-untyped-def]
        """§9: el revisor tiene que poder ver la fila original."""
        assert all(e.evidence_snippet for e in report.extractions)
        assert all(e.source_row for e in report.extractions)


class TestTrickyRows:
    """Las filas concretas donde es fácil equivocarse."""

    def _row(self, report, fragment: str):  # type: ignore[no-untyped-def]
        return next(
            r for r in report.rows if fragment.lower() in str(r.raw.get("title", "")).lower()
        )

    def test_weekly_recurrence_expands_to_four_tuesdays(self, report) -> None:  # type: ignore[no-untyped-def]
        row = self._row(report, "Club de Ajedrez")
        days = sorted(e.starts_at.astimezone(BOGOTA).date() for e in row.extractions)
        assert len(days) == 4
        assert [d.day for d in days] == [7, 14, 21, 28]
        assert all(d.weekday() == 1 for d in days)

    def test_cross_month_range_expands_to_six_dates(self, report) -> None:  # type: ignore[no-untyped-def]
        row = self._row(report, "Semillero de rob")
        days = sorted(e.starts_at.astimezone(BOGOTA).date() for e in row.extractions)
        assert [(d.month, d.day) for d in days] == [
            (6, 23),
            (6, 25),
            (6, 30),
            (7, 2),
            (7, 7),
            (7, 9),
        ]

    def test_only_out_of_month_dates_carry_the_warning(self, report) -> None:  # type: ignore[no-untyped-def]
        """La advertencia es por ACTIVIDAD, no por fila.

        De las 6 fechas solo 3 son de junio; marcar las 6 haría que el panel
        señalara como fuera de mes actividades que sí caen en julio.
        """
        row = self._row(report, "Semillero de rob")
        flagged = [
            e for e in row.extractions if IngestionWarning.OUT_OF_MONTH in e.warnings
        ]
        assert len(flagged) == 3
        assert all(e.starts_at.astimezone(BOGOTA).month == 6 for e in flagged)

    def test_out_of_month_rows_are_not_discarded(self, report) -> None:  # type: ignore[no-untyped-def]
        row = self._row(report, "Semillero de rob")
        assert len(row.extractions) == 6
        assert row.status is not RowStatus.ERROR

    def test_noon_is_not_midnight(self, report) -> None:  # type: ignore[no-untyped-def]
        """'10:00 a.m. a 12:00 m.' → 10:00-12:00, no 10:00-00:00."""
        row = self._row(report, "Aloe Vera")
        first = row.extractions[0]
        assert first.starts_at.astimezone(BOGOTA).hour == 10
        assert first.ends_at is not None
        assert first.ends_at.astimezone(BOGOTA).hour == 12

    def test_non_url_link_is_not_stored_as_url(self, report) -> None:  # type: ignore[no-untyped-def]
        """El caso que el contrato no cubría."""
        row = self._row(report, "Semillero de rob")
        extraction = row.extractions[0]
        assert extraction.registration_url is None
        assert extraction.requires_registration is None
        assert IngestionWarning.REGISTRATION_NOT_A_URL in extraction.warnings
        assert "cúpos completados" in extraction.extra["registration_note"]

    def test_sentinel_room_becomes_null(self, report) -> None:  # type: ignore[no-untyped-def]
        """'No especificado en el documento' → sin sala, sin texto."""
        row = self._row(report, "Club Ambiental")
        extraction = row.extractions[0]
        assert extraction.room_id is None
        assert extraction.room_raw is None

    def test_numbered_rooms_resolve_differently(self, report) -> None:  # type: ignore[no-untyped-def]
        generic = self._row(report, "Herramientas con Inteligencia").extractions[0]
        three = self._row(report, "Análisis de datos").extractions[0]
        four = self._row(report, "NotebookLM").extractions[0]
        assert None not in (generic.room_id, three.room_id, four.room_id)
        assert len({generic.room_id, three.room_id, four.room_id}) == 3

    def test_case_insensitive_rooms_are_the_same(self, report) -> None:  # type: ignore[no-untyped-def]
        """'Taller Infantil' y 'Taller infantil' → la misma sala."""
        upper = self._row(report, "¡A Jugar!").extractions[0]
        lower = self._row(report, "Amasar, jugar").extractions[0]
        assert upper.room_id == lower.room_id
        assert upper.room_id is not None

    def test_grouped_dates_share_an_id(self, report) -> None:  # type: ignore[no-untyped-def]
        """§3.1: para poder editarlas o borrarlas en bloque desde el panel."""
        row = self._row(report, "Haz tu juego con Scratch")
        assert len(row.extractions) == 3
        group_ids = {e.activity_group_id for e in row.extractions}
        assert len(group_ids) == 1
        assert group_ids.pop() is not None

    def test_single_date_has_no_group(self, report) -> None:  # type: ignore[no-untyped-def]
        row = self._row(report, "Taller de collage")
        assert len(row.extractions) == 1
        assert row.extractions[0].activity_group_id is None

    def test_open_ended_age_range(self, report) -> None:  # type: ignore[no-untyped-def]
        row = self._row(report, "Haz tu juego con Scratch")
        extraction = row.extractions[0]
        assert extraction.age_min == 9
        assert extraction.age_max is None


class TestTimezone:
    def test_all_datetimes_are_utc(self, report) -> None:  # type: ignore[no-untyped-def]
        """En la base todo va en UTC (CLAUDE.md §5)."""
        for extraction in report.extractions:
            assert extraction.starts_at.tzinfo is not None
            assert extraction.starts_at.utcoffset() == datetime.now(UTC).utcoffset()

    def test_bogota_offset_is_applied(self, report) -> None:  # type: ignore[no-untyped-def]
        """14:00 en Bogotá son las 19:00 UTC. Colombia no tiene horario de
        verano, así que el desplazamiento es fijo."""
        afternoon = [
            e for e in report.extractions if e.starts_at.astimezone(BOGOTA).hour == 14
        ]
        assert afternoon
        assert all(e.starts_at.hour == 19 for e in afternoon)


class TestResilience:
    """El archivo lo produce un equipo humano cada mes: tiene que aguantar."""

    def _workbook(self, rows: list[Any], sheet_name: str = "Programación") -> Workbook:
        book = Workbook()
        sheet = cast("Worksheet", book.active)
        sheet.title = sheet_name
        for row in rows:
            sheet.append(row)
        return book

    HEADERS = (
        "Título del curso",
        "Descripción",
        "Día(s)",
        "Fecha(s)",
        "Horario",
        "Lugar",
        "Público",
        "Inscripción",
        "Enlace de inscripción",
    )
    DATA = (
        "Taller de prueba",
        "Una descripción",
        "Miércoles",
        "01 de julio",
        "2:00 p.m. a 4:00 p.m.",
        "Taller Infantil",
        "Niñas y niños de 4 a 7 años",
        "No requiere inscripción",
        None,
    )

    def test_headers_on_row_four(self, tmp_path: Path, context: ImportContext) -> None:
        book = self._workbook(
            [
                ("Programación infantil – Julio 2026",),
                (),
                ("una nota suelta",),
                self.HEADERS,
                self.DATA,
            ]
        )
        path = tmp_path / "raro.xlsx"
        book.save(path)
        report = XlsxProgramacionSource().parse(path, context)
        assert report.rows_read == 1
        assert len(report.extractions) == 1

    def test_renamed_sheet_is_recognized_by_headers(
        self, tmp_path: Path, context: ImportContext
    ) -> None:
        """§1: la hoja se reconoce por encabezados, nunca por su nombre."""
        book = self._workbook(
            [("Programación – Julio 2026",), self.HEADERS, self.DATA],
            sheet_name="Hoja renombrada por alguien",
        )
        path = tmp_path / "renombrada.xlsx"
        book.save(path)
        report = XlsxProgramacionSource().parse(path, context)
        assert report.rows_read == 1
        assert report.sheets_skipped == []

    def test_unknown_column_is_kept_not_dropped(
        self, tmp_path: Path, context: ImportContext
    ) -> None:
        """§2: se conserva en `extra` y se reporta."""
        book = self._workbook(
            [
                ("Programación – Julio 2026",),
                (*self.HEADERS, "Cupo máximo"),
                (*self.DATA, "25"),
            ]
        )
        path = tmp_path / "extra.xlsx"
        book.save(path)
        report = XlsxProgramacionSource().parse(path, context)
        assert "Cupo máximo" in report.unknown_columns
        extraction = report.extractions[0]
        assert extraction.extra["unknown_columns"]["Cupo máximo"] == "25"
        assert IngestionWarning.UNKNOWN_COLUMNS in extraction.warnings

    def test_sheet_without_headers_is_skipped_not_fatal(
        self, tmp_path: Path, context: ImportContext
    ) -> None:
        book = Workbook()
        first = cast("Worksheet", book.active)
        first.title = "Notas"
        first.append(["apuntes internos"])
        second = book.create_sheet("Programación")
        second.append(("Programación – Julio 2026",))
        second.append(self.HEADERS)
        second.append(self.DATA)
        path = tmp_path / "mixto.xlsx"
        book.save(path)

        report = XlsxProgramacionSource().parse(path, context)
        assert report.sheets_skipped == ["Notas"]
        assert report.rows_read == 1

    def test_bad_row_does_not_abort_the_file(
        self, tmp_path: Path, context: ImportContext
    ) -> None:
        """La garantía central: un error de fila no tumba la importación."""
        broken = (
            "Fila rota",
            "desc",
            "Lunes",
            "cuando se pueda",  # fecha ininteligible
            "2:00 p.m. a 4:00 p.m.",
            "Taller Infantil",
            "Adultos",
            None,
            None,
        )
        book = self._workbook(
            [("Programación – Julio 2026",), self.HEADERS, self.DATA, broken, self.DATA]
        )
        path = tmp_path / "con-error.xlsx"
        book.save(path)

        report = XlsxProgramacionSource().parse(path, context)
        assert report.rows_read == 3
        assert report.rows_error == 1
        # Las otras dos entraron igual.
        assert len(report.extractions) == 2

    def test_empty_rows_are_ignored(self, tmp_path: Path, context: ImportContext) -> None:
        book = self._workbook(
            [("Programación – Julio 2026",), self.HEADERS, self.DATA, (), (None, None)]
        )
        path = tmp_path / "vacias.xlsx"
        book.save(path)
        assert XlsxProgramacionSource().parse(path, context).rows_read == 1

    def test_missing_title_in_sheet_falls_back_to_parameter(
        self, tmp_path: Path, context: ImportContext
    ) -> None:
        """Sin año en el título se usa el del parámetro, y se marca."""
        book = self._workbook([("Programación formativa",), self.HEADERS, self.DATA])
        path = tmp_path / "sin-mes.xlsx"
        book.save(path)
        report = XlsxProgramacionSource().parse(path, context)
        assert IngestionWarning.MONTH_FROM_PARAMETER in report.extractions[0].warnings
        assert report.extractions[0].confidence < 1.0
