"""`XlsxProgramacionSource`: lee la parrilla mensual del Excel de la Fundación.

Es la **fuente primaria** de programación (ADR 009) y la única determinista:
no pasa por el LLM. Que un modelo reinterprete una celda ya estructurada sería
riesgo sin beneficio.

Garantía central del contrato: **un error de fila nunca aborta el archivo.**
Cada fila se procesa aislada y produce su propio resultado; el operador ve en
el panel qué entró, qué entró con advertencia y qué no entró y por qué.
"""

import json
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import IO, Any
from uuid import uuid4
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from src.application.ingestion.schemas import (
    ActivityExtraction,
    ImportReport,
    IngestionWarning,
    RowResult,
    RowStatus,
)
from src.infrastructure.ingestion.excel.audience import parse_audience
from src.infrastructure.ingestion.excel.dates import (
    DateParseError,
    parse_dates,
    parse_month_year,
    parse_weekdays,
)
from src.infrastructure.ingestion.excel.headers import HeaderMap, find_header_row
from src.infrastructure.ingestion.excel.registration import parse_registration
from src.infrastructure.ingestion.excel.rooms import RoomCatalog, resolve_room
from src.infrastructure.ingestion.excel.times import ParsedTime, TimeParseError, parse_time_range

#: Todo el archivo está expresado en hora de Colombia; en la base se guarda
#: UTC (CLAUDE.md §5). Colombia no tiene horario de verano, así que la
#: conversión es un desplazamiento fijo y no hay fechas ambiguas.
BOGOTA = ZoneInfo("America/Bogota")


@dataclass(frozen=True, slots=True)
class ImportContext:
    """Datos que NO están en el archivo y aporta el operador (§8)."""

    venue_slug: str
    year: int
    month: int
    rooms: RoomCatalog
    file_name: str = "programacion.xlsx"


class XlsxProgramacionSource:
    """Convierte un libro de Excel en `ActivityExtraction`."""

    def parse(self, source: Path | str | IO[bytes], context: ImportContext) -> ImportReport:
        """Parsea desde una ruta o desde un archivo en memoria.

        Acepta ambos porque hay dos caminos de entrada: la CLI trabaja con una
        ruta en disco y el panel recibe la subida del navegador, que no llega a
        tocar el sistema de archivos.
        """
        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            return self._parse_workbook(workbook, context)
        finally:
            workbook.close()

    def _parse_workbook(self, workbook: Any, context: ImportContext) -> ImportReport:
        rows: list[RowResult] = []
        skipped: list[str] = []
        unknown_columns: set[str] = set()

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            raw_rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]

            header = find_header_row(raw_rows)
            if header is None or not header.is_usable:
                # Hoja sin los encabezados esperados: se ignora con aviso, no
                # con error fatal (§1).
                skipped.append(sheet_name)
                continue

            unknown_columns.update(header.unknown)

            # El título de la fila 1 es la ÚNICA fuente del año en el libro.
            title_cell = raw_rows[0][0] if raw_rows and raw_rows[0] else None
            period = parse_month_year(str(title_cell) if title_cell else None)
            if period is None:
                year, month = context.year, context.month
                period_from_parameter = True
            else:
                year, month = period
                period_from_parameter = False

            for offset, raw_row in enumerate(raw_rows[header.row_number :]):
                row_number = header.row_number + offset + 1
                if not self._has_content(raw_row, header):
                    continue
                rows.append(
                    self._parse_row(
                        raw_row,
                        header=header,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        year=year,
                        month=month,
                        period_from_parameter=period_from_parameter,
                        context=context,
                    )
                )

        return ImportReport(
            file_name=context.file_name,
            venue_slug=context.venue_slug,
            rows=rows,
            unknown_columns=sorted(unknown_columns),
            sheets_skipped=skipped,
        )

    @staticmethod
    def _has_content(raw_row: tuple[Any, ...], header: HeaderMap) -> bool:
        """Una fila sin título es una fila vacía de relleno, no un error."""
        title = header.value(raw_row, "title")
        return title is not None and bool(str(title).strip())

    def _parse_row(  # noqa: PLR0913
        self,
        raw_row: tuple[Any, ...],
        *,
        header: HeaderMap,
        sheet_name: str,
        row_number: int,
        year: int,
        month: int,
        period_from_parameter: bool,
        context: ImportContext,
    ) -> RowResult:
        raw_values = {
            field: header.value(raw_row, field) for field in header.columns
        }
        extras = header.extras(raw_row)
        # La fila serializada es el `evidence_snippet`: lo que el revisor ve al
        # lado del JSON estructurado para juzgar si se interpretó bien (§9).
        evidence = json.dumps(
            {k: (str(v) if v is not None else None) for k, v in {**raw_values, **extras}.items()},
            ensure_ascii=False,
        )
        location = f"{sheet_name}!{row_number}"

        def failure(message: str) -> RowResult:
            return RowResult(
                sheet=sheet_name,
                row_number=row_number,
                status=RowStatus.ERROR,
                errors=[message],
                raw=raw_values,
            )

        title = str(raw_values.get("title") or "").strip()
        if not title:
            return failure("Sin título: la fila se descarta (§2)")

        warnings: list[IngestionWarning] = []
        confidence = 1.0
        if period_from_parameter:
            warnings.append(IngestionWarning.MONTH_FROM_PARAMETER)
            confidence -= 0.1

        # --- fechas y horario: si fallan, la fila no es interpretable -------
        try:
            parsed_dates = parse_dates(
                raw_values.get("dates_raw"),
                year=year,
                month=month,
                weekdays_raw=raw_values.get("weekdays_raw"),
            )
        except DateParseError as exc:
            return failure(f"Fechas: {exc}")

        try:
            parsed_time = parse_time_range(raw_values.get("time_raw"))
        except TimeParseError as exc:
            return failure(f"Horario: {exc}")

        if parsed_time.end is None:
            warnings.append(IngestionWarning.NO_END_TIME)

        # --- campos que degradan pero no invalidan --------------------------
        room = resolve_room(raw_values.get("room_raw"), context.rooms)
        warnings.extend(room.warnings)
        confidence -= room.confidence_penalty

        audience = parse_audience(raw_values.get("audience_raw"), sheet_name=sheet_name)
        if audience.from_sheet_name:
            warnings.append(IngestionWarning.AUDIENCE_FROM_SHEET_NAME)
            confidence -= 0.1

        registration = parse_registration(
            raw_values.get("registration_raw"), raw_values.get("registration_url")
        )
        warnings.extend(registration.warnings)

        if extras:
            warnings.append(IngestionWarning.UNKNOWN_COLUMNS)

        # Validación cruzada: el día declarado contra el que cae la fecha (§2).
        declared = parse_weekdays(raw_values.get("weekdays_raw"))
        if declared and any(d.weekday() not in declared for d in parsed_dates.dates):
            warnings.append(IngestionWarning.WEEKDAY_MISMATCH)
            confidence -= 0.1

        extra: dict[str, Any] = {**registration.extra}
        if extras:
            extra["unknown_columns"] = {k: str(v) for k, v in extras.items()}

        # Una fila con N fechas produce N actividades unidas por el mismo
        # `activity_group_id`, para poder editarlas o borrarlas en bloque (§3.1).
        group_id = uuid4() if len(parsed_dates.dates) > 1 else None
        confidence = max(0.0, min(1.0, confidence))
        out_of_month = set(parsed_dates.out_of_month)

        def warnings_for(day: Any) -> list[IngestionWarning]:
            """`out_of_month` se marca por ACTIVIDAD, no por fila.

            «Del 23 de junio al 9 de julio» produce 6 fechas de las que solo 3
            son de junio: marcar las 6 haría que el panel señalara como fuera
            de mes actividades que sí caen en el mes cargado.
            """
            day_warnings = [*warnings]
            if day in out_of_month:
                day_warnings.append(IngestionWarning.OUT_OF_MONTH)
            return list(dict.fromkeys(day_warnings))

        extractions = [
            ActivityExtraction(
                title=title,
                venue_slug=context.venue_slug,
                starts_at=self._to_utc(day, parsed_time, start=True),
                ends_at=self._to_utc(day, parsed_time, start=False),
                description=self._clean(raw_values.get("description")),
                room_id=room.room_id,
                room_raw=room.room_raw,
                recurrence=parsed_dates.recurrence,
                audience=audience.audience,
                age_min=audience.age_min,
                age_max=audience.age_max,
                audience_raw=audience.audience_raw,
                requires_registration=registration.requires_registration,
                registration_url=registration.registration_url,
                activity_group_id=group_id,
                confidence=confidence,
                evidence_snippet=evidence,
                source_row=location,
                warnings=warnings_for(day),
                extra=extra,
            )
            for day in parsed_dates.dates
        ]

        # A nivel de fila sí interesa saber que hubo fechas fuera de mes, para
        # que el resumen de la importación lo refleje.
        row_warnings = list(dict.fromkeys(warnings))
        if out_of_month:
            row_warnings.append(IngestionWarning.OUT_OF_MONTH)

        return RowResult(
            sheet=sheet_name,
            row_number=row_number,
            status=RowStatus.WARNING if row_warnings else RowStatus.OK,
            extractions=extractions,
            warnings=row_warnings,
            raw=raw_values,
        )

    @staticmethod
    def _clean(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _to_utc(day: Any, parsed: ParsedTime, *, start: bool) -> datetime | None:
        clock = parsed.start if start else parsed.end
        if clock is None:
            return None
        local = datetime.combine(day, clock, tzinfo=BOGOTA)
        return local.astimezone(UTC)
